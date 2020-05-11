from openpyxl import load_workbook  #对已经存在的excel进行读写操作
# from openpyxl import workbook #用于新建excel 几乎不用

#openpyxl只能处理.xlsx 文件
#1.打开excel文件
wb = load_workbook("cases.xlsx")  #会创建一个workbook对象，是一个excel文件

#2.定位表单
# ws = wb.active #默认获取第一个激活表单，会创建一个worksheet对象，相当于一个表单
ws = ['ride'] #第二种定位表单的方法 []中是sheet名称

#3.定位单元格
one_cell = ws.cell(row=2,column=2) #会创建一个cell对象，相当于一个单元格
#使用cell对象中的value属性，能获取单元格中的值
# print(one_cell.value)

#方法一：将数据写入指定单元格
#修改单元格中的值
# one_cell.value = "负数与负数相乘1"
#方法二：将数据写入指定单元格
one_cell = ws.cell(row=2,column=2,value="负数与负数相乘") #在定位单元格基础上多了value赋值

#修改后必须保存才会成功
#PermissionError: [Errno 13] Permission denied: 'cases.xlsx'
#报上面这个错说明一定是文件是打开状态没有关闭
wb.save("cases.xlsx")