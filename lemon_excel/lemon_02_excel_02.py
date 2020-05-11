from openpyxl import load_workbook  #对已经存在的excel进行读写操作
# from openpyxl import workbook #用于新建excel 几乎不用

#openpyxl只能处理.xlsx 文件
#1.打开excel文件
wb = load_workbook("cases.xlsx")  #会创建一个workbook对象，是一个excel文件

#2.定位表单
# ws = wb.active #默认获取第一个激活表单，会创建一个worksheet对象，相当于一个表单
ws = wb['ride'] #第二种定位表单的方法

#3.定位单元格
# for row in range(ws.min_row+1, ws.max_row+1 ):  #ws.min_row:单元格最小行数 +1是因为第一行不需要。ws.max_row:单元格最大行数 +1是因为range结尾数取不到
#     for col in range(ws.min_column,ws.max_column+1): #和上面一样这里是获取最小列数和最大列数
#         data = ws.cell(row,col).value #获取值
#         print("值为 {}.类型为 {}".format(data,type(data)))
#ws.iter_rows获取指定某一行数据，min_row=1是最小行号，max_row=1是最大行号，values_only=True是获取表格数据
#使用tuple来转换，返回的是嵌套元组的元组
head_data_tuple = tuple(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
#返回的元组用for循环来循环输出
one_list = []
for one_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)): #这里不指定最大行号，默认取全部行数据
    one_list.append(dict(zip(head_data_tuple, one_tuple))) #zip可以将数据转换为一个个元组，用dict转换为字典

print(one_list[0])
# print(one_list[0]['l_data'])
