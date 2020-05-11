from openpyxl import load_workbook
from lemon_handle_config.handle_config import do_config
class HandleExcel:
    '''
    定义处理Excel的类
    '''

    def __init__(self,filename,sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def get_cases(self):
        '''
        获取所有测试用例
        :return:
        '''
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        head_data_tuple = tuple(ws.iter_rows( max_row=1, values_only=True))[0]
        # print(head_data_tuple)
        one_list = []
        for one_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)):
            one_list.append(dict(zip(head_data_tuple, one_tuple)))
        # print(one_list)
        return one_list
    # def get_case(self,row):
    #     '''
    #     获取指定某一行用例
    #     :return:
    #     '''
    #     return self.get_cases()[row-1]
    def write_result(self, row, actual, result):
        '''
        在指定的行写入数据
        :param row: 行号
        :param actual: 实际结果
        :param result: 用例执行结果（Pass或者Fail）
        :return:
        '''
        other_wb = load_workbook(self.filename)
        if self.sheetname is None:
            other_ws = other_wb.active
        else:
            other_ws = other_wb[self.sheetname]
        if isinstance(row, int) and (2 <= row <= other_ws.max_row):
            other_ws.cell(row=row, column=do_config.get_config_int("excel", "actual_col"), value=actual)
            other_ws.cell(row=row, column=do_config.get_config_int("excel", "result_col"), value=result)
            other_wb.save(self.filename)
            other_wb.close()
        else:
            print("传入的行号有误，行号应为大于1的整数")

if __name__ == '__main__':
    filename = "cases.xlsx"
    do_excel = HandleExcel(filename)
    cases = do_excel.get_cases()
    print(cases)
    do_excel.write_result(2,"fangkun","haha")